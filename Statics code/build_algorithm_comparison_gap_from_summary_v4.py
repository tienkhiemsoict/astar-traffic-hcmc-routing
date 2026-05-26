
import argparse
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
GROUP_FILL = PatternFill("solid", fgColor="D9EAF7")
SUB_FILL = PatternFill("solid", fgColor="EAF3FB")
THIN = Side(style="thin", color="AAB7C4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BUCKETS = ["< 5 km", "5 <= distance < 10 km", ">= 10 km"]
ROWS = [
    ("Dijkstra", "Dijkstra"),
    ("A* Origin", "A* Origin"),
    ("DWA* (c=0.5)", "DWA* base"),
    ("DWA* (c=1.0)", "DWA* c=1.0"),
    ("DWA* (c=1.5)", "DWA* c=1.5"),
    ("DWA* (c=2.0)", "DWA* c=2.0"),
]

def to_float(v):
    try:
        if v is None or v == "":
            return None
        return float(v)
    except Exception:
        return None

def load_summary(path: Path):
    data = {}
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        required = {"bucket", "algorithm", "avg_distance_km", "avg_visited_nodes", "avg_runtime_ms"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"summary.csv thiếu cột: {sorted(missing)}")
        for row in reader:
            bucket = row["bucket"].strip()
            algo = row["algorithm"].strip()
            data[(bucket, algo)] = {
                "distance": to_float(row["avg_distance_km"]),
                "nodes": to_float(row["avg_visited_nodes"]),
                "time": to_float(row["avg_runtime_ms"]),
            }
    return data

def abs_gap(algo_val, base_val):
    if algo_val is None or base_val is None or base_val == 0:
        return None
    return abs((algo_val - base_val) / base_val)

def build_report(summary_data, out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    ws.merge_cells("A1:J1")
    ws["A1"] = "SO SÁNH THEO 3 NHÓM KHOẢNG CÁCH (MỐC A* ORIGIN)"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:A3")
    ws["A2"] = "Thuật toán"
    ws["A2"].fill = GROUP_FILL
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rng, label in [("B2:D2", "< 5 km"), ("E2:G2", "5 <= distance < 10 km"), ("H2:J2", ">= 10 km")]:
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value = label
        cell.fill = GROUP_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    subs = ["Gap", "Node", "Time"] * 3
    for col, label in enumerate(subs, start=2):
        c = ws.cell(3, col, label)
        c.fill = SUB_FILL
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

    block_cols = [(2, "< 5 km"), (5, "5 <= distance < 10 km"), (8, ">= 10 km")]

    for row_idx, (display_name, algo_key) in enumerate(ROWS, start=4):
        ws.cell(row_idx, 1, display_name)
        ws.cell(row_idx, 1).alignment = Alignment(horizontal="left", vertical="center")

        for start_col, bucket in block_cols:
            base = summary_data.get((bucket, "A* Origin"), {})
            current = summary_data.get((bucket, algo_key), {})

            gap_val = abs_gap(current.get("distance"), base.get("distance"))
            node_val = current.get("nodes")
            time_val = current.get("time")

            ws.cell(row_idx, start_col, gap_val)
            ws.cell(row_idx, start_col + 1, node_val)
            ws.cell(row_idx, start_col + 2, time_val)

            ws.cell(row_idx, start_col).number_format = "0.0000%"
            ws.cell(row_idx, start_col + 1).number_format = "0.00"
            ws.cell(row_idx, start_col + 2).number_format = "0.000000"

            ws.cell(row_idx, start_col).alignment = Alignment(horizontal="center")
            ws.cell(row_idx, start_col + 1).alignment = Alignment(horizontal="center")
            ws.cell(row_idx, start_col + 2).alignment = Alignment(horizontal="center")

    for r in range(2, 10):
        for c in range(1, 11):
            ws.cell(r, c).border = BORDER

    for col, width in {"A": 18, "B": 11, "C": 11, "D": 11, "E": 18, "F": 11, "G": 11, "H": 11, "I": 11, "J": 11}.items():
        ws.column_dimensions[col].width = width

    raw = wb.create_sheet("Input_Summary")
    headers = ["bucket", "algorithm", "avg_distance_km", "avg_visited_nodes", "avg_runtime_ms"]
    for idx, h in enumerate(headers, start=1):
        cell = raw.cell(1, idx, h)
        cell.value = h
        cell.fill = TITLE_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    rr = 2
    for bucket in BUCKETS:
        for _, algo_key in ROWS:
            vals = summary_data.get((bucket, algo_key), {})
            raw.cell(rr, 1, bucket)
            raw.cell(rr, 2, algo_key)
            raw.cell(rr, 3, vals.get("distance"))
            raw.cell(rr, 4, vals.get("nodes"))
            raw.cell(rr, 5, vals.get("time"))
            for c in range(1, 6):
                raw.cell(rr, c).border = BORDER
            raw.cell(rr, 3).number_format = "0.000000"
            raw.cell(rr, 4).number_format = "0.00"
            raw.cell(rr, 5).number_format = "0.000000"
            rr += 1

    raw.column_dimensions["A"].width = 24
    raw.column_dimensions["B"].width = 20
    raw.column_dimensions["C"].width = 18
    raw.column_dimensions["D"].width = 18
    raw.column_dimensions["E"].width = 18

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--summary-csv", required=True)
    parser.add_argument("--output", default="algorithm_comparison_gap_v4.xlsx")
    args = parser.parse_args()

    summary_path = Path(args.summary_csv)
    data = load_summary(summary_path)
    build_report(data, Path(args.output))
    print(f"Created: {args.output}")

if __name__ == "__main__":
    main()
