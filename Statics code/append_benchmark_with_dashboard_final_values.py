           # XAY DUNG FILE THONG KE THUC NGHIEM CAC THUAT TOAN TIM DUONG DI

import argparse
import csv
import importlib.util
import inspect
import math
import random
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.table import Table, TableStyleInfo


Coords = Dict[int, Tuple[float, float]]
Adj = Dict[int, List[Tuple[int, float]]]

RAW_SHEET_MAP = {
    "lt5_raw": "distance_lt_5km.csv",
    "from5to10_raw": "distance_5_to_10km.csv",
    "gt10_raw": "distance_gt_10km.csv",
}

TITLE_FILL = PatternFill("solid", fgColor="0F243E")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
THIN_GRAY = Side(style="thin", color="C9D2D8")

ALGO_SPECS = {
    # THUAT TOAN DIJKSTRA
    "dijkstra": {
        "display": "Dijkstra",
        "function": "dijkstra_search",
        "files": ["Dijkstra.py", "dijkstra_algorithm.py"],
        "tokens": ["dijkstra"],
    },
    # THUAT TOAN A*
    "astar": {
        "display": "A* Origin",
        "function": "a_star_search",
        "files": [ "AStarOrigin.py", "A_star_algorithm.py", "a_star_algorithm.py"],
        "tokens": ["astar", "origin"],
    },
    # THUAT TOAN DWA* voi gia tri c=0.5
    "dwa": {
        "display": "DWA* base",
        "function": "dwa_star_search",
        "files": [ "DWAStar.py", "dwa_star_algorithm.py"],
        "tokens": ["dwastar"],
    },
    # THUAT TOAN DWA* voi gia tri c=1.0
    "dwac1": {
        "display": "DWA* c=1.0",
        "function": "dwa_starc1_search",
        "files": [ "DWAStar_with_c1.py"],
        "tokens": ["dwastar", "c1"],
    },
    # THUAT TOAN DWA* voi gia tri c=1.5
    "dwac1.5": {
        "display": "DWA* c=1.5",
        "function": "dwa_starc15_search",
        "files": [ "DWAStar_with_c15.py"],
        "tokens": ["dwastar", "c1.5"],
    },
    # THUAT TOAN DWA* voi gia tri c=2.0
    "dwac2": {
        "display": "DWA* c=2.0",
        "function": "dwa_starc2_search",
        "files": [ "DWAStar_with_c2.py"],
        "tokens": ["dwastar", "c2"],
    },    
   
}


def normalize_name(s: str) -> str:
    return re.sub(r"[^0-9a-z]+", "", s.lower())


def safe_module_name(file_path: Path) -> str:
    return "mod_" + re.sub(r"[^0-9a-zA-Z_]+", "_", file_path.stem)


def load_module_from_file(file_path: Path):
    spec = importlib.util.spec_from_file_location(safe_module_name(file_path), file_path)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load module from file: {file_path}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def collect_python_files(project_dir: Path) -> List[Path]:
    files = [p for p in project_dir.rglob("*.py") if p.is_file()]
    return sorted(files, key=lambda p: (len(str(p)), str(p).lower()))


def find_file_by_override(key: str, overrides: dict, project_dir: Path) -> Path | None:
    if key not in overrides:
        return None
    path = Path(overrides[key])
    if not path.is_absolute():
        path = project_dir / path
    if not path.exists():
        raise FileNotFoundError(f"Override for {key} points to a missing file: {path}")
    return path


def resolve_algorithm(key: str, project_dir: Path, overrides: dict):
    if key not in ALGO_SPECS:
        raise ValueError(f"Unknown algorithm key '{key}'. Supported: {', '.join(ALGO_SPECS.keys())}")

    spec = ALGO_SPECS[key]

    # 1) explicit override
    override_path = find_file_by_override(key, overrides, project_dir)
    if override_path is not None:
        module = load_module_from_file(override_path)
        func = getattr(module, spec["function"], None)
        if func is None:
            raise AttributeError(
                f"Override file {override_path.name} found, but function {spec['function']} not found."
            )
        return {
            "key": key,
            "display": spec["display"],
            "function_name": spec["function"],
            "path": override_path,
            "func": func,
        }

    # 2) exact filename matches (current folder and subfolders)
    all_py_files = collect_python_files(project_dir)
    candidate_names = {name.lower() for name in spec["files"]}
    exact = [p for p in all_py_files if p.name.lower() in candidate_names]
    if exact:
        file_path = exact[0]
        module = load_module_from_file(file_path)
        func = getattr(module, spec["function"], None)
        if func is None:
            raise AttributeError(f"File {file_path.name} found but function {spec['function']} not found.")
        return {
            "key": key,
            "display": spec["display"],
            "function_name": spec["function"],
            "path": file_path,
            "func": func,
        }

    # 3) normalized exact stem match
    candidate_norms = {normalize_name(Path(name).stem) for name in spec["files"]}
    norm_exact = [p for p in all_py_files if normalize_name(p.stem) in candidate_norms]
    if norm_exact:
        file_path = norm_exact[0]
        module = load_module_from_file(file_path)
        func = getattr(module, spec["function"], None)
        if func is None:
            raise AttributeError(f"File {file_path.name} found but function {spec['function']} not found.")
        return {
            "key": key,
            "display": spec["display"],
            "function_name": spec["function"],
            "path": file_path,
            "func": func,
        }

    # 4) fuzzy token match
    tokens = [normalize_name(t) for t in spec.get("tokens", [])]
    fuzzy = []
    for p in all_py_files:
        stem_norm = normalize_name(p.stem)
        if all(token in stem_norm for token in tokens):
            fuzzy.append(p)

    if len(fuzzy) == 1:
        file_path = fuzzy[0]
        module = load_module_from_file(file_path)
        func = getattr(module, spec["function"], None)
        if func is None:
            raise AttributeError(
                f"Fuzzy-matched file {file_path.name} found, but function {spec['function']} not found."
            )
        return {
            "key": key,
            "display": spec["display"],
            "function_name": spec["function"],
            "path": file_path,
            "func": func,
        }

    available = [p.relative_to(project_dir).as_posix() for p in all_py_files]
    if fuzzy:
        raise FileNotFoundError(
            f"Multiple fuzzy matches for algorithm '{key}': {fuzzy}. "
            f"Use --algo-path {key}=relative_path.py to choose one."
        )

    raise FileNotFoundError(
        f"Cannot find source file for algorithm '{key}'. "
        f"Tried exact names {spec['files']}. "
        f"Available .py files under project: {available[:50]}"
    )


def load_project_load_graph(project_dir: Path):
    candidates = [project_dir / "load_graph.py", project_dir / "app.py"]
    for file_path in candidates:
        if file_path.exists():
            module = load_module_from_file(file_path)
            func = getattr(module, "load_graph", None)
            if func is not None:
                return func
    raise FileNotFoundError(f"Cannot find load_graph() in load_graph.py or app.py under {project_dir}")


def call_project_load_graph(load_graph_func, nodes_path: str, edges_path: str, time_slot: int, undirected: bool):
    sig = inspect.signature(load_graph_func)
    params = sig.parameters
    if "directed" in params:
        return load_graph_func(nodes_path=nodes_path, edges_path=edges_path, time_slot=time_slot, directed=not undirected)
    return load_graph_func(nodes_path, edges_path, time_slot)


def slot_to_label(slot: int) -> str:
    total_minutes = (slot - 1) * 30
    return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"


def validate_slot(slot: int) -> int:
    if not 1 <= slot <= 48:
        raise ValueError("time_slot must be between 1 and 48.")
    return slot


def get_time_slot_from_user() -> int:
    print("Choose time slot from 1 to 48:")
    for i in range(1, 49):
        print(f"  {i:>2} -> {slot_to_label(i)}")
    while True:
        raw = input("Enter time slot (1-48): ").strip()
        try:
            return validate_slot(int(raw))
        except Exception:
            print("Invalid value. Please enter an integer from 1 to 48.")


def build_headers(algos: List[dict]) -> List[str]:
    headers = [
        "timestamp",
        "time_slot",
        "time_label",
        "start_node",
        "goal_node",
        "bucket_rule",
        "bucket_distance_km_ref",
    ]
    for idx in range(1, len(algos) + 1):
        headers.extend([
            f"algo{idx}_key",
            f"algo{idx}_name",
            f"algo{idx}_distance_km",
            f"algo{idx}_nodes",
            f"algo{idx}_runtime_ms",
        ])
    return headers


def canonical_pair(start: int, goal: int, directed: bool) -> Tuple[int, int]:
    return (start, goal) if directed else tuple(sorted((start, goal)))


def read_csv_rows(path: Path) -> List[dict]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def existing_pairs_from_output(output_dir: Path, time_slot: int, directed: bool) -> set:
    seen = set()
    for filename in RAW_SHEET_MAP.values():
        path = output_dir / filename
        for row in read_csv_rows(path):
            if str(row.get("time_slot", "")) == str(time_slot):
                try:
                    s = int(row["start_node"])
                    g = int(row["goal_node"])
                    seen.add(canonical_pair(s, g, directed))
                except Exception:
                    pass
    return seen


def check_output_config(output_dir: Path, algos: List[dict]) -> None:
    expected_keys = [a["key"] for a in algos]
    for filename in RAW_SHEET_MAP.values():
        rows = read_csv_rows(output_dir / filename)
        if not rows:
            continue
        first = rows[0]
        current = []
        idx = 1
        while f"algo{idx}_key" in first:
            current.append(first.get(f"algo{idx}_key"))
            idx += 1
        if current != expected_keys:
            raise ValueError(
                f"Output directory already contains a different algorithm set in {filename}: "
                f"{current}. Use a different --output-dir for this algorithm list."
            )


def run_algorithm(func, coords: Coords, adj: Adj, start: int, goal: int) -> dict:
    t0 = time.perf_counter()
    path, distance_m, visited_nodes = func(coords, adj, start, goal)
    runtime_ms = (time.perf_counter() - t0) * 1000.0
    success = path is not None and math.isfinite(distance_m)
    return {
        "success": success,
        "distance_m": float(distance_m) if success else math.inf,
        "distance_km": round(float(distance_m) / 1000.0, 6) if success else None,
        "visited_nodes": int(visited_nodes),
        "runtime_ms": round(float(runtime_ms), 6),
    }


def choose_reference_algo_key(bucket_reference: str, selected_keys: List[str]) -> str:
    if bucket_reference == "auto":
        if "dijkstra" in selected_keys:
            return "dijkstra"
        return selected_keys[0]
    if bucket_reference not in selected_keys:
        raise ValueError(f"--bucket-reference '{bucket_reference}' must be one of selected algorithms: {selected_keys}")
    return bucket_reference


def sample_pair_all_algorithms_success(coords: Coords, adj: Adj, algos: List[dict], seen_pairs: set, directed: bool, rng: random.Random, max_attempts: int = 5000):
    node_ids = list(coords.keys())
    if len(node_ids) < 2:
        raise ValueError("Graph must contain at least 2 nodes.")

    for _ in range(max_attempts):
        start, goal = rng.sample(node_ids, 2)
        key = canonical_pair(start, goal, directed)
        if key in seen_pairs:
            continue

        results = {}
        all_success = True
        for algo in algos:
            result = run_algorithm(algo["func"], coords, adj, start, goal)
            results[algo["key"]] = result
            if not result["success"]:
                all_success = False
                break

        if all_success:
            seen_pairs.add(key)
            return start, goal, results

    raise RuntimeError(
        f"Could not find a new pair where all algorithms succeed after {max_attempts} attempts."
    )


def classify_bucket(reference_distance_km: float):
    if reference_distance_km < 5:
        return "distance_lt_5km.csv", "distance < 5 km"
    if reference_distance_km < 10:
        return "distance_5_to_10km.csv", "5 <= distance < 10 km"
    return "distance_gt_10km.csv", "distance >= 10 km"


def build_output_row(time_slot: int, start: int, goal: int, results_by_key: dict, algos: List[dict], reference_distance_km: float, bucket_rule: str) -> dict:
    row = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "time_slot": time_slot,
        "time_label": slot_to_label(time_slot),
        "start_node": start,
        "goal_node": goal,
        "bucket_rule": bucket_rule,
        "bucket_distance_km_ref": reference_distance_km,
    }
    for idx, algo in enumerate(algos, start=1):
        result = results_by_key[algo["key"]]
        row[f"algo{idx}_key"] = algo["key"]
        row[f"algo{idx}_name"] = algo["display"]
        row[f"algo{idx}_distance_km"] = result["distance_km"]
        row[f"algo{idx}_nodes"] = result["visited_nodes"]
        row[f"algo{idx}_runtime_ms"] = result["runtime_ms"]
    return row


def append_row(csv_path: Path, row: dict, headers: List[str]) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not csv_path.exists() or csv_path.stat().st_size == 0
    try:
        with csv_path.open("a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            if write_header:
                writer.writeheader()
            writer.writerow(row)
    except PermissionError as e:
        raise PermissionError(f"Cannot write to {csv_path}. Close the file in Excel/VS Code and run again.") from e


def style_title_row(ws, end_col: int, title: str):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1)
    cell.value = title
    cell.fill = TITLE_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def apply_header_style(cell):
    cell.fill = HEADER_FILL
    cell.font = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(top=THIN_GRAY, bottom=THIN_GRAY)


def apply_metric_style(cell):
    cell.number_format = "0.000000"
    cell.alignment = Alignment(horizontal="right")


def write_raw_sheet(ws, title: str, headers: List[str], rows: List[dict]):
    style_title_row(ws, len(headers), title)
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(2, c, header)
        apply_header_style(cell)
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx, row.get(header))
    ws.freeze_panes = "A3"

    for idx, header in enumerate(headers, start=1):
        col = get_column_letter(idx)
        if header == "timestamp" or header == "bucket_rule" or header.endswith("_name"):
            width = 22
        elif header in {"start_node", "goal_node", "time_slot"} or header.endswith("_nodes"):
            width = 12
        elif header.endswith("_distance_km") or header.endswith("_runtime_ms") or header == "bucket_distance_km_ref":
            width = 16
        else:
            width = 14
        ws.column_dimensions[col].width = width

    if rows:
        end_row = len(rows) + 2
        end_col_letter = get_column_letter(len(headers))
        table = Table(displayName=f"tbl_{ws.title}", ref=f"A2:{end_col_letter}{end_row}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)




def _to_float(value):
    try:
        if value is None or value == "":
            return None
        return float(value)
    except Exception:
        return None


def summarize_rows(rows: List[dict], algos: List[dict]) -> List[dict]:
    summary = []
    bucket_defs = [
        ("< 5 km", rows.get("lt5_raw", [])),
        ("5 <= distance < 10 km", rows.get("from5to10_raw", [])),
        (">= 10 km", rows.get("gt10_raw", [])),
    ]
    for bucket_label, bucket_rows in bucket_defs:
        for idx, algo in enumerate(algos, start=1):
            dist_vals = []
            node_vals = []
            runtime_vals = []
            for row in bucket_rows:
                d = _to_float(row.get(f"algo{idx}_distance_km"))
                n = _to_float(row.get(f"algo{idx}_nodes"))
                t = _to_float(row.get(f"algo{idx}_runtime_ms"))
                if d is not None:
                    dist_vals.append(d)
                if n is not None:
                    node_vals.append(n)
                if t is not None:
                    runtime_vals.append(t)
            summary.append({
                "bucket": bucket_label,
                "algorithm": algo["display"],
                "avg_distance_km": round(sum(dist_vals) / len(dist_vals), 6) if dist_vals else None,
                "avg_visited_nodes": round(sum(node_vals) / len(node_vals), 6) if node_vals else None,
                "avg_runtime_ms": round(sum(runtime_vals) / len(runtime_vals), 6) if runtime_vals else None,
            })
    return summary




def write_summary_csv(output_dir: Path, algos: List[dict]) -> Path:
    raw_rows = {}
    for sheet_name, csv_name in RAW_SHEET_MAP.items():
        raw_rows[sheet_name] = read_csv_rows(output_dir / csv_name)

    summary_rows = summarize_rows(raw_rows, algos)
    summary_path = output_dir / "summary.csv"
    with summary_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["bucket", "algorithm", "avg_distance_km", "avg_visited_nodes", "avg_runtime_ms"],
        )
        writer.writeheader()
        writer.writerows(summary_rows)
    return summary_path


def create_dashboard(workbook_path: Path, output_dir: Path, algos: List[dict], headers: List[str]):
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    raw_rows = {}
    for sheet_name, csv_name in RAW_SHEET_MAP.items():
        raw_rows[sheet_name] = read_csv_rows(output_dir / csv_name)

    meta = wb.create_sheet("Metadata")
    style_title_row(meta, 4, "Benchmark Configuration")
    meta_headers = ["Field", "Value", "Field", "Value"]
    for idx, header in enumerate(meta_headers, start=1):
        apply_header_style(meta.cell(2, idx, header))
    rows = [
        ("Generated at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Output dir", str(output_dir)),
        ("Algorithm count", len(algos), "Workbook", workbook_path.name),
    ]
    for algo_idx, algo in enumerate(algos, start=1):
        rows.append((f"Algo {algo_idx}", algo["display"], "Source file", algo["path"].name))
    for r_idx, row in enumerate(rows, start=3):
        for c_idx, value in enumerate(row, start=1):
            meta.cell(r_idx, c_idx, value)
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 26
    meta.column_dimensions["C"].width = 20
    meta.column_dimensions["D"].width = 30

    summary = wb.create_sheet("Summary")
    style_title_row(summary, 5, "Benchmark Summary by Distance Bucket")
    summary_headers = ["Bucket", "Algorithm", "Avg Distance (km)", "Avg Visited Nodes", "Avg Runtime (ms)"]
    for c, header in enumerate(summary_headers, start=1):
        apply_header_style(summary.cell(2, c, header))

    summary_rows = summarize_rows(raw_rows, algos)
    row_idx = 3
    for item in summary_rows:
        summary.cell(row_idx, 1, item["bucket"])
        summary.cell(row_idx, 2, item["algorithm"])
        summary.cell(row_idx, 3, item["avg_distance_km"])
        summary.cell(row_idx, 4, item["avg_visited_nodes"])
        summary.cell(row_idx, 5, item["avg_runtime_ms"])
        for c in range(1, 6):
            cell = summary.cell(row_idx, c)
            cell.border = Border(bottom=THIN_GRAY)
            if c >= 3:
                apply_metric_style(cell)
        row_idx += 1

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 22
    summary.column_dimensions["C"].width = 20
    summary.column_dimensions["D"].width = 20
    summary.column_dimensions["E"].width = 18
    summary.freeze_panes = "A3"

    for sheet_name, csv_name in RAW_SHEET_MAP.items():
        ws = wb.create_sheet(sheet_name)
        write_raw_sheet(ws, f"Raw data - {csv_name}", headers, raw_rows[sheet_name])

    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)


def parse_algo_path_pairs(values: List[str]) -> dict:
    mapping = {}
    for item in values or []:
        if "=" not in item:
            raise ValueError(f"Invalid --algo-path value '{item}'. Use key=relative_or_absolute_path.py")
        key, path = item.split("=", 1)
        key = key.strip()
        path = path.strip()
        if not key or not path:
            raise ValueError(f"Invalid --algo-path value '{item}'.")
        mapping[key] = path
    return mapping


def parse_args():
    parser = argparse.ArgumentParser(description="Compare multiple algorithms at once, append results to distance-bucket CSV files, and refresh Excel dashboard.")
    parser.add_argument("--nodes", type=str, default="data/nodes.csv")
    parser.add_argument("--edges", type=str, default="data/edges.csv")
    parser.add_argument("--time-slot", type=int, default=None, help="1..48. If omitted, prompt for input.")
    parser.add_argument("--trials", type=int, default=1)
    parser.add_argument("--seed", type=int, default=None)
    parser.add_argument("--undirected", action="store_true")
    parser.add_argument("--algorithms", nargs="+", default=list(ALGO_SPECS.keys()),
                        help="Algorithm keys to compare. Default compares all available algorithms.")
    parser.add_argument("--bucket-reference", default="auto",
                        help="Selected algorithm key used for bucket classification. Default: auto (Dijkstra if selected, else first).")
    parser.add_argument("--output-dir", type=str, default=None,
                        help="Folder to store CSV files. Default uses selected algorithm keys.")
    parser.add_argument("--workbook", type=str, default=None,
                        help="Excel dashboard path. Default inside output dir.")
    parser.add_argument("--allow-duplicate-pairs", action="store_true",
                        help="Allow repeated start-goal pairs already stored in existing CSV files.")
    parser.add_argument("--algo-path", nargs="*", default=[],
                        help="Optional overrides like dijkstra=algos/MyDijkstra.py astar=algos/MyAStar.py")
    return parser.parse_args()


def main():
    args = parse_args()
    project_dir = Path.cwd()

    selected_keys = args.algorithms
    if len(selected_keys) < 2:
        raise ValueError("Please provide at least 2 algorithms.")
    if len(set(selected_keys)) != len(selected_keys):
        raise ValueError("Algorithm keys must be distinct.")

    overrides = parse_algo_path_pairs(args.algo_path)
    algos = [resolve_algorithm(key, project_dir, overrides) for key in selected_keys]
    headers = build_headers(algos)

    output_dir = Path(args.output_dir) if args.output_dir else Path(f"distance_bucket_results__{'_'.join(selected_keys)}")
    workbook_path = Path(args.workbook) if args.workbook else output_dir / "benchmark_dashboard_values.xlsx"

    check_output_config(output_dir, algos)

    selected_slot = get_time_slot_from_user() if args.time_slot is None else validate_slot(args.time_slot)

    load_graph_func = load_project_load_graph(project_dir)
    graph_result = call_project_load_graph(load_graph_func, args.nodes, args.edges, selected_slot, args.undirected)
    if len(graph_result) < 2:
        raise ValueError("load_graph() must return at least coords and adj.")
    coords, adj = graph_result[0], graph_result[1]

    ref_key = choose_reference_algo_key(args.bucket_reference, selected_keys)
    rng = random.Random(args.seed)
    seen_pairs = set() if args.allow_duplicate_pairs else existing_pairs_from_output(output_dir, selected_slot, directed=not args.undirected)

    for trial_index in range(1, args.trials + 1):
        start, goal, results = sample_pair_all_algorithms_success(
            coords, adj, algos, seen_pairs, directed=not args.undirected, rng=rng
        )

        ref_distance_km = results[ref_key]["distance_km"]
        csv_name, bucket_rule = classify_bucket(ref_distance_km)
        row = build_output_row(selected_slot, start, goal, results, algos, ref_distance_km, bucket_rule)
        append_row(output_dir / csv_name, row, headers)
        print(
            f"[{trial_index}/{args.trials}] pair=({start},{goal}) bucket='{bucket_rule}' ref={ref_key} distance={ref_distance_km} km (all algorithms succeeded)"
        )

    create_dashboard(workbook_path, output_dir, algos, headers)
    summary_csv_path = write_summary_csv(output_dir, algos)
    print("\nDone.")
    print(f"Output dir:  {output_dir}")
    print(f"Workbook:    {workbook_path}")
    print(f"Summary CSV: {summary_csv_path}")


if __name__ == "__main__":
    main()

# HUONG DAN SU DUNG:
# 1) Dat file nay vao thu muc chua code cua ban (cung cap tren duong dan tuong doi den file load_graph.py cua ban)
# 2) Chay file nay, no se tu dong tim cac ham thuat toan trong code cua ban (theo quy tac dat ten va token), chay tung thuat toan tren nhung cap start-goal ngau nhien ma tat ca thuat toan deu thanh cong, va luu ket qua vao cac file CSV phu hop voi quy tac bucket theo khoang cach.
# 3) No se tu dong tao file Excel dashboard va cap nhat cac gia tri vao day, cung cap cac bang tong hop va cho phep ban xem du lieu thuc nghiem chi tiet trong Excel. Ban co the mo file Excel sau moi lan chay de xem ket qua moi duoc them vao. Neu file Excel dang mo, dong no lai truoc khi chay lai de cap nhat du lieu.